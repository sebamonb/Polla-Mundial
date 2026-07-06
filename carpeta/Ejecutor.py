import re
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

# --- CONFIGURACIÓN ---
MASTER_FILE = "participantes3.xlsx"   # <-- ajusta si el nombre cambia
CARPETA = "."
HEADER_MASTER = False        # el maestro NO tiene fila de encabezado
HEADER_PARTICIPANTE = False  # el excel individual NO tiene fila de encabezado


def normalizar(texto):
    texto = str(texto).strip().lower()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto)
                    if unicodedata.category(c) != 'Mn')
    return re.sub(r'[^a-z0-9]', '', texto)


def encontrar_archivo_participante(nombre, carpeta):
    nombre_norm = normalizar(nombre)
    archivos = [a for a in Path(carpeta).glob("*.xlsx") if not a.name.startswith('~$')
                and a.resolve() != Path(MASTER_FILE).resolve()]
    for a in archivos:
        if normalizar(a.stem) == nombre_norm:
            return a
    for a in archivos:
        base = normalizar(a.stem)
        if nombre_norm in base or base in nombre_norm:
            return a
    return None


def leer_partidos_maestro(ruta_maestro):
    wb = load_workbook(ruta_maestro, data_only=True)
    ws = wb.active
    fila_inicio = 2 if HEADER_MASTER else 1

    participantes = {}
    for fila in ws.iter_rows(min_row=fila_inicio, values_only=True):
        nombre = fila[0]
        if nombre is None or str(nombre).strip() == "":
            continue
        valores = fila[1:]
        partidos = []
        for i in range(0, len(valores) - 1, 2):
            local, visita = valores[i], valores[i + 1]
            if local is None and visita is None:
                continue
            partidos.append((local, visita))
        participantes[str(nombre).strip()] = partidos
    return participantes


def ya_fueron_agregados(ws, partidos_nuevos):
    """Revisa si las ultimas N filas del excel individual ya coinciden
    exactamente con los partidos que queremos agregar (para no duplicar)."""
    n = len(partidos_nuevos)
    if ws.max_row < n:
        return False

    fila_inicio = ws.max_row - n + 1
    filas_actuales = []
    for fila in ws.iter_rows(min_row=fila_inicio, max_row=ws.max_row, values_only=True):
        filas_actuales.append((fila[1], fila[2]))

    return filas_actuales == partidos_nuevos


def actualizar_participante(ruta_participante, partidos_nuevos):
    wb = load_workbook(ruta_participante)
    ws = wb.active

    if ya_fueron_agregados(ws, partidos_nuevos):
        return 0  # ya estaban cargados, no duplicar

    fila_siguiente = ws.max_row + 1
    for local, visita in partidos_nuevos:
        ws.cell(row=fila_siguiente, column=2, value=local)
        ws.cell(row=fila_siguiente, column=3, value=visita)
        fila_siguiente += 1

    wb.save(ruta_participante)
    return len(partidos_nuevos)


def main():
    participantes = leer_partidos_maestro(MASTER_FILE)
    print(f"Participantes leídos del maestro: {len(participantes)}\n")

    for nombre, partidos in participantes.items():
        archivo = encontrar_archivo_participante(nombre, CARPETA)
        if archivo is None:
            print(f"⚠️  No se encontró excel para '{nombre}'")
            continue
        agregados = actualizar_participante(archivo, partidos)
        if agregados:
            print(f"✅ {nombre}: {agregados} partido(s) agregado(s) en {archivo.name}")
        else:
            print(f"— {nombre}: ya estaban cargados, no se dupl. ({archivo.name})")


if __name__ == "__main__":
    main()
