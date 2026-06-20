import requests
from openpyxl import load_workbook

# =========================
# MAPA ESPAÑOL → ESPN
# =========================

mapa = {
    "Mexico": "Mexico",
    "Sudafrica": "South Africa",
    "Corea del Sur": "South Korea",
    "Republica Checa": "Czechia",
    "Canada": "Canada",
    "Bosnia y Herzegovina": "Bosnia-Herzegovina",
    "Estados Unidos": "United States",
    "Paraguay": "Paraguay",
    "Qatar": "Qatar",
    "Suiza": "Switzerland",
    "Brasil": "Brazil",
    "Marruecos": "Morocco",
    "Haiti": "Haiti",
    "Escocia": "Scotland",
    "Australia": "Australia",
    "Turquia": "Turkey",
    "Alemania": "Germany",
    "Ecuador":"Ecuador",
    "Curacao": "Curaçao",
    "Paises Bajos": "Netherlands",
    "Japon": "Japan",
    "Costa de Marfil": "Ivory Coast",
    "Suecia": "Sweden",
    "Tunisia": "Tunisia",
    "Espana": "Spain",
    "Cabo Verde": "Cape Verde",
    "Belgica": "Belgium",
    "Egipto": "Egypt",
    "Arabia Saudita": "Saudi Arabia",
    "Uruguay": "Uruguay",
    "Iran": "Iran",
    "Nueva Zelanda": "New Zealand",
    "Francia": "France",
    "Iraq": "Iraq",
    "Noruega": "Norway",
    "Argentina": "Argentina",
    "Argelia": "Algeria",
    "Austria": "Austria",
    "Jordania": "Jordan",
    "Portugal": "Portugal",
    "Croacia": "Croatia",
    "RD Congo": "Congo DR",
    "Inglaterra": "England",
    "Ghana": "Ghana",
    "Panama": "Panama",
    "Uzbekistan": "Uzbekistan",
    "Colombia": "Colombia"
}

# =========================
# API ESPN
# =========================

URL = "https://site.api.espn.com/apis/site/v2/sports/soccer/fifa.world/scoreboard"
r = requests.get(URL, timeout=30)
data = r.json()

events = data.get("events", [])
print(f"Eventos recibidos: {len(events)}")
for e in events:
    comp = e["competitions"][0]
    print(f"  {e['name']} - completed: {comp['status']['type']['completed']} - status: {comp['status']['type']['description']}")
print("=== NOMBRES EXACTOS DE LA API ===")
for e in events:
    comp = e["competitions"][0]
    for c in comp["competitors"]:
        print(repr(c["team"]["displayName"]))
# =========================
# EXCEL (OPENPYXL)
# =========================

ARCHIVO = "resultados.xlsx"

wb = load_workbook(ARCHIVO)
ws = wb.active

# =========================
# RECORRER PARTIDOS
# =========================

for event in events:

    comp = event["competitions"][0]

    if not comp["status"]["type"]["completed"]:
        continue

    competitors = comp["competitors"]

    home_team = None
    away_team = None
    home_score = None
    away_score = None

    for c in competitors:
        team = c["team"]["displayName"]
        score = c["score"]

        if c["homeAway"] == "home":
            home_team = team
            home_score = score
        else:
            away_team = team
            away_score = score

    # =========================
    # BUSCAR EN EXCEL
    # =========================

    for i in range(2, ws.max_row + 1):

        excel_home = ws.cell(i, 1).value
        excel_away = ws.cell(i, 4).value
        flag = ws.cell(i, 8).value

        if str(flag).strip() == "1":
            continue

        excel_home_en = mapa.get(excel_home, excel_home)
        excel_away_en = mapa.get(excel_away, excel_away)

        if (excel_home_en == home_team and excel_away_en == away_team):

            ws.cell(i, 2).value = home_score
            ws.cell(i, 3).value = away_score
            ws.cell(i, 8).value = 1

        elif (excel_home_en == away_team and excel_away_en == home_team):

            ws.cell(i, 2).value = away_score
            ws.cell(i, 3).value = home_score
            ws.cell(i, 8).value = 1

# =========================
# GUARDAR
# =========================

wb.save(ARCHIVO)

print("Excel actualizado correctamente")
